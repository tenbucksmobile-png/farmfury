#!/usr/bin/env python3
"""Automated Excel reconciliation processor.

Usage:
    python recon_processor.py RAW_DATA.xlsx

Produces RECON_OUTPUT.xlsx in the same directory as the input file.
"""

import os
import sys
from collections import defaultdict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def get_row_type(desc: str) -> str:
    s = str(desc).strip()
    if s.startswith('SDH'): return 'SDH'
    if s.startswith('SRH'): return 'SRH'
    if s.startswith('SC'):  return 'SC'
    if s.startswith('SI'):  return 'SI'
    return 'OTHER'


def is_populated(val) -> bool:
    v = str(val).strip().lower()
    return v not in ('', '0', '0.0', '0.00', 'nan', 'none', 'nat', '#n/a')


def is_blank_or_zero(val) -> bool:
    return not is_populated(val)


# ---------------------------------------------------------------------------
# Column-name globals (set in main after reading the file)
# ---------------------------------------------------------------------------
COL_DESC = COL_INV_EXC = COL_INV_INC = None
COL_CR_EXCL = COL_CR_INCL = None
COL_SERIAL = COL_ADD_REF = COL_ADD_REF1 = None


def sort_group(rows):
    """
    Sort an SI group so that each SDH row is immediately followed by
    its matching SC row(s) (same Serial Number = paired/batched together).
    Remaining SC rows that had no SDH pair in this group are appended after.
    SRH rows go last.
    """
    si    = [(i, r) for i, r in rows if r['_rtype'] == 'SI']
    sdh   = [(i, r) for i, r in rows if r['_rtype'] == 'SDH']
    sc    = [(i, r) for i, r in rows if r['_rtype'] == 'SC']
    srh   = [(i, r) for i, r in rows if r['_rtype'] == 'SRH']
    other = [(i, r) for i, r in rows if r['_rtype'] == 'OTHER']

    # Map SN → list of SC rows within this group
    sc_by_sn = defaultdict(list)
    for item in sc:
        sn = item[1][COL_SERIAL].strip()
        sc_by_sn[sn].append(item)

    used_sc_idx = set()
    body = []
    for sdh_item in sdh:
        body.append(sdh_item)
        sn = sdh_item[1][COL_SERIAL].strip()
        for sc_item in sc_by_sn.get(sn, []):
            if sc_item[0] not in used_sc_idx:
                body.append(sc_item)
                used_sc_idx.add(sc_item[0])

    # SC rows with no matching SDH in this group
    for sc_item in sc:
        if sc_item[0] not in used_sc_idx:
            body.append(sc_item)

    return si + body + srh + other


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    global COL_DESC, COL_INV_EXC, COL_INV_INC
    global COL_CR_EXCL, COL_CR_INCL
    global COL_SERIAL, COL_ADD_REF, COL_ADD_REF1

    if len(sys.argv) < 2:
        print("Usage: python recon_processor.py <input_file.xlsx>")
        sys.exit(1)

    input_path = sys.argv[1]

    if not os.path.exists(input_path):
        print(f"Error: File '{input_path}' not found.")
        sys.exit(1)

    output_dir  = os.path.dirname(os.path.abspath(input_path))
    output_path = os.path.join(output_dir, "RECON_OUTPUT.xlsx")

    # ------------------------------------------------------------------
    # Read
    # ------------------------------------------------------------------
    df = pd.read_excel(input_path, dtype=str)
    df = df.fillna('').replace('nan', '').replace('NaT', '').replace('None', '')

    if len(df) == 0:
        print("Warning: Input file has no data rows. Writing empty output.")
        wb = Workbook()
        ws = wb.active
        ws.title = "Recon"
        wb.save(output_path)
        print(f"Output file: {output_path}")
        return

    orig_cols = list(df.columns[:19])

    COL_DESC     = orig_cols[3]
    COL_INV_EXC  = orig_cols[8]
    COL_INV_INC  = orig_cols[9]
    COL_CR_EXCL  = orig_cols[10]
    COL_CR_INCL  = orig_cols[11]
    COL_ADD_REF  = orig_cols[15]
    COL_SERIAL   = orig_cols[16]
    COL_ADD_REF1 = orig_cols[18]

    df['_rtype'] = df[COL_DESC].apply(get_row_type)

    # ------------------------------------------------------------------
    # Read original row formatting from the raw file (font + fill)
    # so the output preserves whatever colours were in the source.
    # pandas idx 0 == Excel row 2, idx 1 == row 3, etc.
    # ------------------------------------------------------------------
    from openpyxl import load_workbook as _lw

    raw_row_colors: dict[int, tuple] = {}   # pandas_idx → (font_rgb, fill_rgb|None)
    try:
        _rwb = _lw(input_path)
        _rws = _rwb.active
        for _pidx, _xl_row in enumerate(_rws.iter_rows(min_row=2)):
            _font_rgb = 'FF000000'   # default black
            _fill_rgb = None
            # Determine row colour from the Description cell (col 4, index 3).
            # Scanning all cells would pick up stray FFFF0000 on empty columns
            # (e.g. the CR-amount columns) and wrongly paint black rows red.
            try:
                _dcell = _xl_row[3] if len(_xl_row) > 3 else _xl_row[0]
                if (_dcell.font and _dcell.font.color
                        and _dcell.font.color.type == 'rgb'):
                    _c = _dcell.font.color.rgb
                    if _c and _c.upper() != '00000000':
                        _font_rgb = _c
            except Exception:
                pass
            for _cell in _xl_row:
                try:
                    if (_cell.fill and _cell.fill.fill_type == 'solid'
                            and _cell.fill.fgColor
                            and _cell.fill.fgColor.type == 'rgb'):
                        _c = _cell.fill.fgColor.rgb
                        if _c and _c.upper() != '00000000':
                            _fill_rgb = _c
                            break
                except Exception:
                    pass
            raw_row_colors[_pidx] = (_font_rgb, _fill_rgb)
        _rwb.close()
    except Exception as e:
        print(f"Warning: could not read source formatting ({e}). Using default black.")

    # ------------------------------------------------------------------
    # Build linkage maps
    # ------------------------------------------------------------------

    # sn_to_si: serial number → SI number (from SDH rows)
    sn_to_si: dict[str, str] = {}
    for _, row in df[df['_rtype'] == 'SDH'].iterrows():
        sn = row[COL_SERIAL].strip()
        si = row[COL_ADD_REF1].strip()
        if sn and si:
            sn_to_si[sn] = si

    # SC serial numbers (for unmatched-SDH detection)
    sc_serials: set[str] = set()
    for _, row in df[df['_rtype'] == 'SC'].iterrows():
        sn = row[COL_SERIAL].strip()
        if sn:
            sc_serials.add(sn)

    # Duplicate SN detection
    sn_counts: dict[str, int] = defaultdict(int)
    for _, row in df.iterrows():
        sn = row[COL_SERIAL].strip()
        if sn:
            sn_counts[sn] += 1
    duplicate_sns: set[str] = {sn for sn, c in sn_counts.items() if c > 1}

    # ------------------------------------------------------------------
    # Group rows by SI number
    # resolved_sc_si tracks which SI each SC row was linked to,
    # so we can write it into the AR1 column in the output.
    # ------------------------------------------------------------------

    si_groups: dict[str, list] = defaultdict(list)
    orphaned_rows: list = []
    unlinked_sc_indices: set[int] = set()
    resolved_sc_si: dict[int, str] = {}   # orig_idx → resolved SI number

    for idx, row in df.iterrows():
        rtype = row['_rtype']
        desc  = row[COL_DESC].strip()

        if rtype == 'SI':
            si_groups[desc].append((idx, row))

        elif rtype == 'SDH':
            si = row[COL_ADD_REF1].strip()
            if si:
                si_groups[si].append((idx, row))
            else:
                orphaned_rows.append((idx, row))

        elif rtype == 'SC':
            sn = row[COL_SERIAL].strip()
            si = sn_to_si.get(sn, '')
            if si:
                si_groups[si].append((idx, row))
                resolved_sc_si[idx] = si          # remember resolved SI for AR1 write
            else:
                unlinked_sc_indices.add(idx)
                si_groups['__UNLINKED_SC__'].append((idx, row))

        elif rtype == 'SRH':
            si = row[COL_ADD_REF1].strip()
            if not si:
                si = row[COL_ADD_REF].strip()
            if si:
                si_groups[si].append((idx, row))
            else:
                sn = row[COL_SERIAL].strip()
                linked_si = sn_to_si.get(sn, '')
                if linked_si:
                    si_groups[linked_si].append((idx, row))
                else:
                    orphaned_rows.append((idx, row))

        else:
            orphaned_rows.append((idx, row))

    # ------------------------------------------------------------------
    # Build flat output list
    # ------------------------------------------------------------------

    output_rows: list = []

    if orphaned_rows:
        output_rows.extend(orphaned_rows)
        output_rows.append(None)

    for si_num, rows in si_groups.items():
        if si_num == '__UNLINKED_SC__':
            continue
        output_rows.extend(sort_group(rows))
        output_rows.append(None)

    if '__UNLINKED_SC__' in si_groups:
        output_rows.extend(si_groups['__UNLINKED_SC__'])
        output_rows.append(None)

    # ------------------------------------------------------------------
    # Write output workbook
    # ------------------------------------------------------------------

    wb = Workbook()
    ws = wb.active
    ws.title = "Recon"

    col_widths = [14, 12, 22, 22, 12, 38, 10, 22, 14, 14, 14, 14, 12, 22, 14, 22, 20, 10, 22]
    for ci, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    GREY_FILL   = PatternFill(fill_type='solid', fgColor='FFD9D9D9')
    NO_FILL     = PatternFill(fill_type=None)
    COLOR_BLACK = 'FF000000'

    for ci, hdr in enumerate(orig_cols, start=1):
        cell = ws.cell(row=1, column=ci, value=hdr)
        cell.font = Font(name='Arial', size=10, bold=True, color=COLOR_BLACK)
        cell.fill = GREY_FILL

    # Numeric columns (1-based): I=9 J=10 K=11 L=12 M=13
    NUM_COLS = {9, 10, 11, 12, 13}

    si_vat_added  = 0
    sc_vat_added  = 0
    unmatched_sdh = 0
    output_n      = 0
    excel_row     = 2

    for item in output_rows:
        if item is None:
            excel_row += 1
            continue

        orig_idx, row = item
        rtype = row['_rtype']
        sn    = row[COL_SERIAL].strip()

        # ── VAT decisions ────────────────────────────────────────────────
        # SI rows: VAT in Inv Amt Inc (col J) if Inv Amt Exc populated
        add_inv_vat = (
            rtype == 'SI'
            and is_populated(row[COL_INV_EXC])
            and is_blank_or_zero(row[COL_INV_INC])
        )

        # SC rows: prefer CR Amt Excl (col K) → CR Amt Incl (col L)
        #          fall back to Inv Amt Exc (col I) → Inv Amt Inc (col J)
        add_cr_vat  = (
            rtype == 'SC'
            and is_populated(row[COL_CR_EXCL])
            and is_blank_or_zero(row[COL_CR_INCL])
        )
        add_inv_vat_sc = (
            rtype == 'SC'
            and not add_cr_vat
            and is_populated(row[COL_INV_EXC])
            and is_blank_or_zero(row[COL_INV_INC])
        )
        add_sc_vat = add_cr_vat or add_inv_vat_sc

        # Use the original row colours from the raw file
        _orig_font, _orig_fill = raw_row_colors.get(orig_idx, ('FF000000', None))
        row_font = Font(name='Arial', size=10, color=_orig_font)
        row_fill = (PatternFill(fill_type='solid', fgColor=_orig_fill)
                    if _orig_fill else NO_FILL)

        # ── Write 19 cells ───────────────────────────────────────────────
        for ci, col_name in enumerate(orig_cols, start=1):
            cell = ws.cell(row=excel_row, column=ci)

            # VAT formula injection
            if rtype == 'SI' and add_inv_vat and ci == 10:
                # Inv Amt Inc = Inv Amt Exc * 1.15
                cell.value = f'=I{excel_row}*1.15'

            elif rtype == 'SC' and add_cr_vat and ci == 12:
                # CR Amt Incl = CR Amt Excl * 1.15
                cell.value = f'=K{excel_row}*1.15'

            elif rtype == 'SC' and add_inv_vat_sc and ci == 10:
                # SC row with amount in Inv Amt Exc: Inv Amt Inc = Inv Amt Exc * 1.15
                cell.value = f'=I{excel_row}*1.15'

            elif rtype == 'SC' and ci == 19:
                # AR1 column: write the resolved SI number for linked SC rows
                resolved = resolved_sc_si.get(orig_idx, '')
                raw_val  = row[col_name]
                cell.value = resolved if resolved else (raw_val or None)

            else:
                raw = row[col_name]
                if ci in NUM_COLS:
                    try:
                        v = str(raw).strip()
                        cell.value = float(v) if v else None
                    except ValueError:
                        cell.value = raw or None
                elif ci == 1:
                    if raw:
                        try:
                            cell.value = pd.to_datetime(raw).to_pydatetime()
                        except Exception:
                            cell.value = raw
                    else:
                        cell.value = None
                else:
                    cell.value = raw or None

            cell.font = row_font
            cell.fill = row_fill

            if ci in NUM_COLS:
                cell.number_format = '#,##0.00'
            elif ci == 1:
                cell.number_format = 'DD/MM/YYYY'

        # ── Stats ────────────────────────────────────────────────────────
        if add_inv_vat:
            si_vat_added += 1
        if add_sc_vat:
            sc_vat_added += 1
        if rtype == 'SDH' and sn not in sc_serials:
            unmatched_sdh += 1
        output_n += 1
        excel_row += 1

    wb.save(output_path)

    # ------------------------------------------------------------------
    # Summary
    # ------------------------------------------------------------------
    si_group_count = len([k for k in si_groups if k != '__UNLINKED_SC__'])

    print("=== RECON PROCESSOR SUMMARY ===")
    print(f"Input rows processed:       {len(df)}")
    print(f"SI groups found:            {si_group_count}")
    print(f"SC rows with VAT added:     {sc_vat_added}")
    print(f"SI rows with VAT added:     {si_vat_added}")
    print(f"Duplicate Serial Numbers:   {len(duplicate_sns)}  (unique SNs appearing 2+ times)")
    print(f"Unmatched SDH rows:         {unmatched_sdh}   (SDH rows with SN not in any SC)")
    print(f"Unlinked SC rows:           {len(unlinked_sc_indices)}    (SC rows whose SN not found in any SDH)")
    print(f"Output rows written:        {output_n}")
    print(f"Output file:                {output_path}")
    print("================================")


if __name__ == '__main__':
    main()
